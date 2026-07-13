"""Generate the flagship demo workbook (GRAPH-D1.1).

Deterministic (seeded) so the workbook is reproducible. Builds a hand-designed
flagship cluster around Z_PRICE_ENGINE plus disconnected, believable filler
clusters, emits demo-data/flagship.xlsx per docs/excel-schema.md, and writes the
validation artifact docs/flagship-blast-radius.md from the SAME in-memory model
(so the summary matches the workbook by construction).

Run:
    ./backend/.venv/bin/python backend/scripts/generate_flagship_dataset.py
"""

from __future__ import annotations

import random
from collections import Counter, defaultdict, deque
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[2]
XLSX_PATH = REPO / "demo-data" / "flagship.xlsx"
BLAST_MD_PATH = REPO / "docs" / "flagship-blast-radius.md"

SEED = 20260713
TARGET_NODES = 214
TARGET_DIRECT_EDGES = 386

# ---------------------------------------------------------------------------
# Model
# ---------------------------------------------------------------------------
# name -> {"label", "description", "business_meaning"}
nodes: dict[str, dict[str, str]] = {}
# distinct (source, type, target) direct edges
edges: list[tuple[str, str, str]] = []
_edge_set: set[tuple[str, str, str]] = set()


def add_node(name: str, label: str, description: str = "", business_meaning: str = "") -> str:
    if name not in nodes:
        nodes[name] = {
            "label": label,
            "description": description,
            "business_meaning": business_meaning,
        }
    else:
        # Fill detail if a previously-implicit node is later described.
        if description and not nodes[name]["description"]:
            nodes[name]["description"] = description
        if business_meaning and not nodes[name]["business_meaning"]:
            nodes[name]["business_meaning"] = business_meaning
    return name


def add_edge(source: str, rel_type: str, target: str) -> None:
    key = (source, rel_type, target)
    if key not in _edge_set:
        _edge_set.add(key)
        edges.append(key)


# ---------------------------------------------------------------------------
# Flagship cluster — hand-designed, exact (per the ticket)
# ---------------------------------------------------------------------------
def build_flagship() -> None:
    add_node("Z_PRICE_ENGINE", "Program",
             "Core pricing calculation program.",
             "Prices every sales order line. Owned by SD; any change needs pricing-team sign-off.")
    add_node("Z_REPORT_MARGIN", "Program",
             "Gross-margin report over the pricing log.",
             "Controlling reads this weekly. Built on the price log Z_PRICE_ENGINE writes.")
    add_node("Z_REPORT_COMPLIANCE", "Program",
             "Pricing compliance extract for audit.",
             "Feeds the quarterly external audit pack off the price audit trail.")
    add_node("Z_LEGACY_REBATE_CALC", "Program",
             "Legacy rebate calculation - origin unknown.",
             "Nobody left owns this. Kept in case month-end ever needs it. Not scheduled.")

    add_node("Z_FM_VALIDATE_PRICE", "FunctionModule",
             "Validates a computed price against business rules.",
             "Called by the pricing engine before commit; logs any failures.")
    add_node("Z_FM_CURRENCY_CONVERT", "FunctionModule",
             "Currency conversion helper.",
             "Converts to document currency using the daily rate table.")
    add_node("Z_FM_TAX_LOOKUP", "FunctionModule",
             "Tax code / rate lookup.",
             "Returns the tax rate for a country + tax code combination.")

    add_node("ZBADI_PRICE_OVERRIDE", "BAdI",
             "Customer-specific price override hook.",
             "Key-account override logic, implemented inside the pricing engine.")

    add_node("ZJOB_NIGHTLY_PRICING", "Job",
             "Nightly pricing recalculation.",
             "Runs the pricing engine across the open order book overnight.")
    add_node("ZJOB_MONTH_END_CLOSE", "Job",
             "Month-end close pricing step.",
             "Re-prices open items as part of the FI month-end close.")
    add_node("ZJOB_WEEKLY_REPORTING", "Job",
             "Weekly reporting batch.",
             "Runs the margin report every Monday 06:00.")
    add_node("ZJOB_QUARTERLY_AUDIT", "Job",
             "Quarterly compliance extract job.",
             "Produces the audit pack for external auditors each quarter.")

    add_node("TR_2024_0847", "Transport", "Pricing engine hotfix - rounding correction.")
    add_node("TR_2025_0112", "Transport", "Pricing engine change - new discount tier.")

    # Z_PRICE_ENGINE direct edges
    for t in ("ZTABLE_PRICING", "ZTABLE_CONDITIONS", "ZTABLE_CUSTOMER_DISCOUNTS"):
        add_node(t, "Table")
        add_edge("Z_PRICE_ENGINE", "READS", t)
    for t in ("ZTABLE_PRICE_LOG", "ZTABLE_PRICE_AUDIT"):
        add_node(t, "Table")
        add_edge("Z_PRICE_ENGINE", "WRITES", t)
    for f in ("Z_FM_VALIDATE_PRICE", "Z_FM_CURRENCY_CONVERT", "Z_FM_TAX_LOOKUP"):
        add_edge("Z_PRICE_ENGINE", "CALLS", f)

    # FM behaviour (Z_FM_VALIDATE_PRICE WRITES gives the engine its inferred write)
    add_node("ZTABLE_VALIDATION_ERRORS", "Table")
    add_edge("Z_FM_VALIDATE_PRICE", "WRITES", "ZTABLE_VALIDATION_ERRORS")
    add_node("ZTABLE_EXCHANGE_RATES", "Table")
    add_edge("Z_FM_CURRENCY_CONVERT", "READS", "ZTABLE_EXCHANGE_RATES")
    add_node("ZTABLE_TAX_RATES", "Table")
    add_edge("Z_FM_TAX_LOOKUP", "READS", "ZTABLE_TAX_RATES")

    # Downstream consumers of the engine's output tables (adds depth)
    add_edge("Z_REPORT_MARGIN", "READS", "ZTABLE_PRICE_LOG")
    add_edge("Z_REPORT_COMPLIANCE", "READS", "ZTABLE_PRICE_AUDIT")

    # BAdI, Jobs, Transports
    add_edge("Z_PRICE_ENGINE", "IMPLEMENTS", "ZBADI_PRICE_OVERRIDE")
    add_edge("ZJOB_NIGHTLY_PRICING", "EXECUTES", "Z_PRICE_ENGINE")
    add_edge("ZJOB_MONTH_END_CLOSE", "EXECUTES", "Z_PRICE_ENGINE")
    add_edge("ZJOB_WEEKLY_REPORTING", "EXECUTES", "Z_REPORT_MARGIN")
    add_edge("ZJOB_QUARTERLY_AUDIT", "EXECUTES", "Z_REPORT_COMPLIANCE")
    add_edge("TR_2024_0847", "CHANGED", "Z_PRICE_ENGINE")
    add_edge("TR_2025_0112", "CHANGED", "Z_PRICE_ENGINE")


# ---------------------------------------------------------------------------
# Filler clusters — believable, disconnected from the flagship cluster
# ---------------------------------------------------------------------------
THEMES = [
    ("STOCK", "inventory valuation", "Logistics"),
    ("INVOICE", "customer invoicing", "Finance"),
    ("PROCURE", "procurement", "Purchasing"),
    ("PAYROLL", "payroll", "HR"),
    ("GL", "general ledger", "Finance"),
    ("SALESORD", "sales order", "Sales"),
    ("SHIPPING", "outbound delivery", "Logistics"),
    ("WAREHOUSE", "warehouse movement", "Logistics"),
    ("ASSET", "asset accounting", "Finance"),
    ("VENDOR", "vendor master", "Purchasing"),
    ("MATERIAL", "material master", "Logistics"),
    ("BILLING", "billing run", "Finance"),
]
PROGRAMS_PER = [4, 4, 4, 4, 4, 4, 4, 4, 4, 3, 3, 3]   # 45
FMS_PER = [3, 3, 3, 3, 3, 3, 3, 3, 3, 2, 2, 2]        # 33
TABLES_PER = [8, 8, 8, 7, 7, 7, 7, 7, 7, 7, 7, 7]     # 87
JOBS_PER = [2, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 2]       # 14
BADIS_PER = [1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0]      # 5
TRANS_PER = [1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0]      # 8

PROG_STEMS = ["POSTING", "EXTRACT", "RECONCILE", "UPDATE", "LOAD", "ADJUST",
              "REVALUATE", "CLEANUP", "MIGRATE", "ARCHIVE", "SPLIT", "MERGE"]
FM_STEMS = ["GET_DATA", "VALIDATE", "CONVERT", "LOOKUP", "CALCULATE", "FORMAT", "CHECK"]
TABLE_STEMS = ["HEADER", "ITEM", "LOG", "HISTORY", "STAGING", "MASTER", "INDEX",
               "STATUS", "DELTA", "ARCHIVE"]
JOB_STEMS = ["NIGHTLY", "WEEKLY", "DAILY", "MONTHEND"]
BADI_STEMS = ["OVERRIDE", "EXTEND", "HOOK"]
CAVEATS = [
    "Do not run during period close.",
    "Undocumented; predates the current team.",
    "Superseded by S/4 standard but still scheduled.",
    "Original owner left the company.",
    "Runs under a dedicated batch user.",
]

_tr_seq = 500  # global transport sequence (avoids collision with flagship TRs)


def _next_transport_id() -> str:
    global _tr_seq
    _tr_seq += 7
    year = 2023 + (_tr_seq % 3)
    return f"TR_{year}_{_tr_seq:04d}"


def build_filler(rng: random.Random) -> None:
    theme_programs: list[list[str]] = []
    theme_fms: list[list[str]] = []
    theme_tables: list[list[str]] = []

    for ti, (prefix, label, dept) in enumerate(THEMES):
        progs, fms, tables, jobs, badis, transports = [], [], [], [], [], []

        for i in range(PROGRAMS_PER[ti]):
            name = f"Z_{prefix}_{PROG_STEMS[i]}"
            add_node(name, "Program",
                     f"{label.capitalize()} {PROG_STEMS[i].lower()} program.",
                     f"{dept} owns this. {rng.choice(CAVEATS)}")
            progs.append(name)
        for i in range(FMS_PER[ti]):
            name = f"Z_FM_{prefix}_{FM_STEMS[i]}"
            add_node(name, "FunctionModule",
                     f"{label} {FM_STEMS[i].replace('_', ' ').lower()}.",
                     f"Called from the {label} programs.")
            fms.append(name)
        for i in range(TABLES_PER[ti]):
            # Tables are implicit (no Tables sheet) - created via edges below.
            tables.append(f"ZTABLE_{prefix}_{TABLE_STEMS[i]}")
        for i in range(JOBS_PER[ti]):
            name = f"ZJOB_{prefix}_{JOB_STEMS[i]}"
            add_node(name, "Job",
                     f"{JOB_STEMS[i].capitalize()} {label} batch.",
                     f"Scheduled in {dept}'s batch window.")
            jobs.append(name)
        for i in range(BADIS_PER[ti]):
            name = f"ZBADI_{prefix}_{BADI_STEMS[i]}"
            add_node(name, "BAdI",
                     f"{label.capitalize()} enhancement hook.",
                     f"Custom {label} logic for {dept}.")
            badis.append(name)
        for _ in range(TRANS_PER[ti]):
            name = _next_transport_id()
            add_node(name, "Transport", f"{label.capitalize()} change request.")
            transports.append(name)

        theme_programs.append(progs)
        theme_fms.append(fms)
        theme_tables.append(tables)

        # Wire edges within the theme (no cross-theme, no flagship links).
        for p in progs:
            for t in rng.sample(tables, min(len(tables), rng.randint(2, 3))):
                add_node(t, "Table")
                add_edge(p, "READS", t)
            for t in rng.sample(tables, rng.randint(1, 2)):
                add_node(t, "Table")
                add_edge(p, "WRITES", t)
            for f in rng.sample(fms, min(len(fms), rng.randint(1, 2))):
                add_edge(p, "CALLS", f)
        for f in fms:
            t = rng.choice(tables)
            add_node(t, "Table")
            add_edge(f, "READS", t)
            if rng.random() < 0.5:
                t2 = rng.choice(tables)
                add_node(t2, "Table")
                add_edge(f, "WRITES", t2)
        for j in jobs:
            for p in rng.sample(progs, min(len(progs), rng.randint(1, 2))):
                add_edge(j, "EXECUTES", p)
        for b in badis:
            add_edge(rng.choice(progs), "IMPLEMENTS", b)
        for tr in transports:
            pool = progs + fms + tables
            for o in rng.sample(pool, min(len(pool), rng.randint(2, 4))):
                add_node(o, "Table" if o in tables else nodes[o]["label"])
                add_edge(tr, "CHANGED", o)

    # Ensure every table is referenced at least once (so it exists on parse).
    referenced = {tg for _, ty, tg in edges if ty in ("READS", "WRITES")}
    for ti, tables in enumerate(theme_tables):
        for t in tables:
            if t not in referenced:
                add_node(t, "Table")
                add_edge(theme_programs[ti][0], "READS", t)
                referenced.add(t)

    # Deterministically pad READS up to the exact direct-edge target.
    for ti, progs in enumerate(theme_programs):
        for p in progs:
            for t in theme_tables[ti]:
                if len(edges) >= TARGET_DIRECT_EDGES:
                    return
                if (p, "READS", t) not in _edge_set and (p, "WRITES", t) not in _edge_set:
                    add_node(t, "Table")
                    add_edge(p, "READS", t)


# ---------------------------------------------------------------------------
# Inference + blast radius
# ---------------------------------------------------------------------------
def compute_inferred() -> list[tuple[str, str, str]]:
    writes_by_src: dict[str, set[str]] = defaultdict(set)
    for s, ty, tg in edges:
        if ty == "WRITES":
            writes_by_src[s].add(tg)
    inferred: set[tuple[str, str, str]] = set()
    for a, ty, b in edges:
        if ty != "CALLS":
            continue
        for t in writes_by_src.get(b, ()):
            if (a, "WRITES", t) not in _edge_set:
                inferred.add((a, "WRITES", t))
    return sorted(inferred)


def compute_blast_radius(start: str, inferred: list[tuple[str, str, str]]):
    """Forward impact traversal from `start` (see docs/flagship-blast-radius.md
    for the documented rule). Returns {node: (hop, confidence, via)}."""
    out_writes: dict[str, set[tuple[str, str]]] = defaultdict(set)
    out_calls: dict[str, set[str]] = defaultdict(set)
    out_impl: dict[str, set[str]] = defaultdict(set)
    table_readers: dict[str, set[str]] = defaultdict(set)
    prog_execs: dict[str, set[str]] = defaultdict(set)
    for s, ty, tg in edges:
        if ty == "WRITES":
            out_writes[s].add((tg, "direct"))
        elif ty == "CALLS":
            out_calls[s].add(tg)
        elif ty == "IMPLEMENTS":
            out_impl[s].add(tg)
        elif ty == "READS":
            table_readers[tg].add(s)
        elif ty == "EXECUTES":
            prog_execs[tg].add(s)
    for s, ty, tg in inferred:
        if ty == "WRITES":
            out_writes[s].add((tg, "inferred"))

    result: dict[str, tuple[int, str, str]] = {start: (0, "direct", "start")}
    q: deque[str] = deque([start])
    while q:
        n = q.popleft()
        hop = result[n][0]
        lab = nodes[n]["label"]
        neighbors: list[tuple[str, str, str]] = []
        if lab in ("Program", "FunctionModule"):
            for tbl, conf in out_writes.get(n, ()):
                neighbors.append((tbl, conf, f"{n} -{'inferred ' if conf == 'inferred' else ''}WRITES-> {tbl}"))
            for f in out_calls.get(n, ()):
                neighbors.append((f, "direct", f"{n} CALLS {f}"))
            for bd in out_impl.get(n, ()):
                neighbors.append((bd, "direct", f"{n} IMPLEMENTS {bd}"))
            if lab == "Program":
                for j in prog_execs.get(n, ()):
                    neighbors.append((j, "direct", f"{j} EXECUTES {n}"))
        elif lab == "Table":
            for r in table_readers.get(n, ()):
                neighbors.append((r, "direct", f"{r} READS {n}"))
        # Job and BAdI are terminal.
        for m, conf, via in sorted(neighbors):
            if m not in result:
                result[m] = (hop + 1, conf, via)
                q.append(m)
    del result[start]
    return result


# ---------------------------------------------------------------------------
# Emit workbook
# ---------------------------------------------------------------------------
def _by_label(label: str) -> list[str]:
    return [n for n, d in nodes.items() if d["label"] == label]


def _targets(source: str, rel_type: str) -> str:
    return ", ".join(tg for s, ty, tg in edges if s == source and ty == rel_type)


def write_workbook() -> None:
    reads = writes = calls = None  # noqa: F841 (documented cell semantics)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Programs")
    ws.append(["Program", "Description", "BusinessMeaning", "Reads", "Writes", "Calls"])
    for n in _by_label("Program"):
        d = nodes[n]
        ws.append([n, d["description"], d["business_meaning"],
                   _targets(n, "READS"), _targets(n, "WRITES"), _targets(n, "CALLS")])

    ws = wb.create_sheet("FunctionModules")
    ws.append(["FunctionModule", "Description", "BusinessMeaning", "Reads", "Writes", "Calls"])
    for n in _by_label("FunctionModule"):
        d = nodes[n]
        ws.append([n, d["description"], d["business_meaning"],
                   _targets(n, "READS"), _targets(n, "WRITES"), _targets(n, "CALLS")])

    ws = wb.create_sheet("BAdIs")
    ws.append(["BAdI", "Description", "BusinessMeaning", "ImplementedBy"])
    implementer = {tg: s for s, ty, tg in edges if ty == "IMPLEMENTS"}
    for n in _by_label("BAdI"):
        d = nodes[n]
        ws.append([n, d["description"], d["business_meaning"], implementer.get(n, "")])

    ws = wb.create_sheet("Jobs")
    ws.append(["Job", "Description", "BusinessMeaning", "Executes"])
    for n in _by_label("Job"):
        d = nodes[n]
        ws.append([n, d["description"], d["business_meaning"], _targets(n, "EXECUTES")])

    ws = wb.create_sheet("Transports")
    ws.append(["Transport", "Description", "Changed"])
    for n in _by_label("Transport"):
        d = nodes[n]
        ws.append([n, d["description"], _targets(n, "CHANGED")])

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)


# ---------------------------------------------------------------------------
# Emit blast-radius.md
# ---------------------------------------------------------------------------
FLAGSHIP = "Z_PRICE_ENGINE"
FLAGSHIP_CLUSTER = [
    "Z_PRICE_ENGINE", "Z_REPORT_MARGIN", "Z_REPORT_COMPLIANCE", "Z_LEGACY_REBATE_CALC",
    "Z_FM_VALIDATE_PRICE", "Z_FM_CURRENCY_CONVERT", "Z_FM_TAX_LOOKUP",
    "ZBADI_PRICE_OVERRIDE",
    "ZJOB_NIGHTLY_PRICING", "ZJOB_MONTH_END_CLOSE", "ZJOB_WEEKLY_REPORTING", "ZJOB_QUARTERLY_AUDIT",
    "TR_2024_0847", "TR_2025_0112",
    "ZTABLE_PRICING", "ZTABLE_CONDITIONS", "ZTABLE_CUSTOMER_DISCOUNTS",
    "ZTABLE_PRICE_LOG", "ZTABLE_PRICE_AUDIT", "ZTABLE_VALIDATION_ERRORS",
    "ZTABLE_EXCHANGE_RATES", "ZTABLE_TAX_RATES",
]


def write_blast_radius_md(inferred: list[tuple[str, str, str]]) -> None:
    node_counts = Counter(d["label"] for d in nodes.values())
    edge_counts = Counter(ty for _, ty, _ in edges)
    label_order = ["Program", "Table", "FunctionModule", "BAdI", "Job", "Transport"]
    rel_order = ["READS", "WRITES", "CALLS", "IMPLEMENTS", "EXECUTES", "CHANGED"]

    radius = compute_blast_radius(FLAGSHIP, inferred)
    ordered = sorted(radius.items(), key=lambda kv: (kv[1][0], nodes[kv[0]]["label"], kv[0]))
    radius_types = sorted({nodes[n]["label"] for n in radius})
    n_inferred = sum(1 for _, (_, conf, _) in radius.items() if conf == "inferred")

    flagship_direct = [(s, ty, tg) for (s, ty, tg) in edges
                       if s in FLAGSHIP_CLUSTER and tg in FLAGSHIP_CLUSTER]
    flagship_inferred = [(s, ty, tg) for (s, ty, tg) in inferred if s in FLAGSHIP_CLUSTER]

    lines: list[str] = []
    lines.append("# Flagship Blast Radius — validation artifact (GRAPH-D1.1)\n")
    lines.append("_Auto-generated by `backend/scripts/generate_flagship_dataset.py` from the same")
    lines.append("model that emits `demo-data/flagship.xlsx`. A human validates the demo scenario")
    lines.append("by reading this against the spreadsheet._\n")

    lines.append("## Totals\n")
    lines.append(f"- **Total nodes:** {len(nodes)}  (target 214)")
    lines.append(f"- **Total direct edges:** {len(edges)}  (target 386)")
    lines.append(f"- **Inferred edges (added by importer, not in the workbook):** {len(inferred)}\n")

    lines.append("### Nodes per label\n")
    lines.append("| Label | Count |")
    lines.append("|---|---|")
    for lab in label_order:
        lines.append(f"| {lab} | {node_counts.get(lab, 0)} |")
    lines.append(f"| **Total** | **{sum(node_counts.values())}** |\n")

    lines.append("### Direct edges per relationship type\n")
    lines.append("| Type | Count |")
    lines.append("|---|---|")
    for rel in rel_order:
        lines.append(f"| {rel} | {edge_counts.get(rel, 0)} |")
    lines.append(f"| **Total** | **{sum(edge_counts.values())}** |\n")

    lines.append("---\n")
    lines.append("## Blast-radius traversal rule\n")
    lines.append("Blast radius of a node **X** = X's forward impact footprint. Starting from X,")
    lines.append("follow, transitively (shortest hop wins):\n")
    lines.append("- from a **Program/FunctionModule**: its outbound `WRITES` (direct *and* inferred),")
    lines.append("  outbound `CALLS`, and (Programs only) outbound `IMPLEMENTS`;")
    lines.append("- from a **Program**: also the **Jobs that `EXECUTE` it** (they fail if it breaks);")
    lines.append("- from a **Table**: the **Programs/FMs that `READ` it** (they consume X's output);")
    lines.append("- **Jobs** and **BAdIs** are terminal.\n")
    lines.append("Outbound `READS` (a node's *inputs*) are **not** traversed — an input is not")
    lines.append("downstream of X. This is why `ZTABLE_EXCHANGE_RATES` / `ZTABLE_TAX_RATES` (inputs")
    lines.append("to the called FMs) are in the graph but **not** in the blast radius.\n")

    lines.append(f"## `{FLAGSHIP}` blast radius — {len(radius)} downstream nodes, "
                 f"{len(radius_types)} node types, {len(radius) - n_inferred} direct + {n_inferred} inferred\n")
    lines.append("| # | Downstream node | Type | Hop | Confidence | Via |")
    lines.append("|---|---|---|---|---|---|")
    for i, (name, (hop, conf, via)) in enumerate(ordered, 1):
        lines.append(f"| {i} | {name} | {nodes[name]['label']} | {hop} | {conf} | {via} |")
    lines.append("")
    lines.append(f"Node types in radius: {', '.join(radius_types)}.\n")

    lines.append("Deepest chains (3 hops):")
    lines.append("- `Z_PRICE_ENGINE` →(WRITES)→ `ZTABLE_PRICE_LOG` →(READS by)→ `Z_REPORT_MARGIN`"
                 " →(EXECUTES by)→ `ZJOB_WEEKLY_REPORTING`")
    lines.append("- `Z_PRICE_ENGINE` →(WRITES)→ `ZTABLE_PRICE_AUDIT` →(READS by)→ `Z_REPORT_COMPLIANCE`"
                 " →(EXECUTES by)→ `ZJOB_QUARTERLY_AUDIT`\n")

    lines.append("---\n")
    lines.append("## Flagship cluster — full edge listing (trace against the workbook)\n")
    lines.append(f"Direct edges ({len(flagship_direct)}):\n")
    for s, ty, tg in flagship_direct:
        lines.append(f"- `{s}` -[{ty}]-> `{tg}`")
    lines.append("")
    lines.append(f"Inferred edges ({len(flagship_inferred)}):\n")
    for s, ty, tg in flagship_inferred:
        lines.append(f"- `{s}` -[{ty} *(inferred)*]-> `{tg}`  "
                     f"— transitive: {s}→Z_FM_VALIDATE_PRICE→{tg}")
    lines.append("")
    lines.append("**Sparse node (empty-state demo):** `Z_LEGACY_REBATE_CALC` — a Program with"
                 " zero edges in or out. Not in any blast radius.\n")

    BLAST_MD_PATH.write_text("\n".join(lines))


# ---------------------------------------------------------------------------
def main() -> None:
    rng = random.Random(SEED)
    build_flagship()
    build_filler(rng)
    inferred = compute_inferred()
    write_workbook()
    write_blast_radius_md(inferred)

    node_counts = Counter(d["label"] for d in nodes.values())
    edge_counts = Counter(ty for _, ty, _ in edges)
    radius = compute_blast_radius(FLAGSHIP, inferred)

    print(f"Nodes: {len(nodes)} (target {TARGET_NODES})   "
          f"Direct edges: {len(edges)} (target {TARGET_DIRECT_EDGES})   "
          f"Inferred: {len(inferred)}")
    print("Nodes per label:", dict(node_counts))
    print("Direct edges per type:", dict(edge_counts))
    print(f"{FLAGSHIP} blast radius: {len(radius)} nodes across "
          f"{len({nodes[n]['label'] for n in radius})} types")
    print(f"Wrote {XLSX_PATH.relative_to(REPO)} and {BLAST_MD_PATH.relative_to(REPO)}")

    assert abs(len(nodes) - TARGET_NODES) <= 5, "node total out of tolerance"
    assert abs(len(edges) - TARGET_DIRECT_EDGES) <= 5, "edge total out of tolerance"


if __name__ == "__main__":
    main()
