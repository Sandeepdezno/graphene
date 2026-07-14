"""Excel importer: parse a workbook conforming to docs/excel-schema.md into
domain nodes + relationships.

No Neo4j / persistence here - this adapter only returns domain objects
(GRAPH-D1.2). Persistence is GRAPH-D1.5's job.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.importer_ports import SourceImporter
from domain.models import (
    BAdI,
    Confidence,
    FunctionModule,
    ImportResult,
    Job,
    NodeModel,
    Program,
    Relationship,
    RelationshipType,
    Table,
    Transport,
)
from .inference import infer_transitive_writes


class ImporterValidationError(Exception):
    """Raised when the workbook is missing a required sheet or column."""


NODE_CLASSES: dict[str, type[NodeModel]] = {
    "Program": Program,
    "Table": Table,
    "FunctionModule": FunctionModule,
    "BAdI": BAdI,
    "Job": Job,
    "Transport": Transport,
}

# sheet -> required columns (first column is the node-name column)
REQUIRED_COLUMNS: dict[str, list[str]] = {
    "Programs": ["Program", "Description", "BusinessMeaning", "Reads", "Writes", "Calls"],
    "FunctionModules": [
        "FunctionModule", "Description", "BusinessMeaning", "Reads", "Writes", "Calls",
    ],
    "BAdIs": ["BAdI", "Description", "BusinessMeaning", "ImplementedBy"],
    "Jobs": ["Job", "Description", "BusinessMeaning", "Executes"],
    "Transports": ["Transport", "Description", "Changed"],
}


def _split(cell: str) -> list[str]:
    """Comma-separated cell -> trimmed non-empty names."""
    return [part.strip() for part in cell.split(",") if part.strip()]


class ExcelImporter(SourceImporter):
    """Parses the flagship workbook contract into domain objects."""

    @staticmethod
    def _load(content: bytes):
        return load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    def validate(self, content: bytes) -> None:
        """Structure-only check (sheets + columns). Raises ImporterValidationError.

        Cheap enough to run synchronously in the request handler so a malformed
        upload gets a 422 before any job is created.
        """
        wb = self._load(content)
        try:
            self._validate(wb)
        finally:
            wb.close()

    def parse_direct(
        self, content: bytes
    ) -> tuple[list[NodeModel], list[Relationship]]:
        """Parse the workbook into nodes + direct (explicit) edges only."""
        wb = self._load(content)
        try:
            self._validate(wb)
            now = datetime.now(timezone.utc)
            registry: dict[str, NodeModel] = {}
            direct: list[Relationship] = []
            seen: set[tuple[str, str, str]] = set()

            self._parse_code_sheet(wb["Programs"], "Programs", "Program", registry, direct, seen, now)
            self._parse_code_sheet(
                wb["FunctionModules"], "FunctionModules", "FunctionModule",
                registry, direct, seen, now,
            )
            self._parse_badis(wb["BAdIs"], registry, direct, seen, now)
            self._parse_jobs(wb["Jobs"], registry, direct, seen, now)
            self._parse_transports(wb["Transports"], registry, direct, seen, now)

            return list(registry.values()), direct
        finally:
            wb.close()

    def infer(self, direct: list[Relationship]) -> list[Relationship]:
        """Apply the one inference rule to a set of direct edges."""
        return infer_transitive_writes(direct, datetime.now(timezone.utc))

    def parse(self, content: bytes) -> ImportResult:
        nodes, direct = self.parse_direct(content)
        inferred = self.infer(direct)
        return ImportResult(nodes=nodes, relationships=direct + inferred)

    # -- validation -------------------------------------------------------
    def _validate(self, wb) -> None:
        for sheet, columns in REQUIRED_COLUMNS.items():
            if sheet not in wb.sheetnames:
                raise ImporterValidationError(f"Missing required sheet: '{sheet}'")
            header = self._header(wb[sheet])
            for column in columns:
                if column not in header:
                    raise ImporterValidationError(
                        f"Sheet '{sheet}' is missing required column: '{column}'"
                    )

    @staticmethod
    def _header(ws: Worksheet) -> list[str]:
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [str(v).strip() if v is not None else "" for v in row]
        return []

    # -- node/edge helpers ------------------------------------------------
    @staticmethod
    def _ensure_node(
        registry: dict[str, NodeModel],
        name: str,
        label: str,
        now: datetime,
        description: str = "",
        business_meaning: str = "",
    ) -> None:
        if name not in registry:
            registry[name] = NODE_CLASSES[label](
                name=name,
                description=description,
                business_meaning=business_meaning,
                source="excel_import",
                confidence=Confidence.DIRECT,
                last_updated=now,
            )
        else:
            node = registry[name]
            # Fill detail if the node was first seen implicitly (empty).
            if description and not node.description:
                node.description = description
            if business_meaning and not node.business_meaning:
                node.business_meaning = business_meaning

    @staticmethod
    def _add_edge(
        direct: list[Relationship],
        seen: set[tuple[str, str, str]],
        source_name: str,
        rel_type: RelationshipType,
        target_name: str,
        derived_from: str,
        now: datetime,
    ) -> None:
        key = (source_name, rel_type.value, target_name)
        if key in seen:  # dedupe identical direct triples
            return
        seen.add(key)
        direct.append(
            Relationship(
                source_name=source_name,
                target_name=target_name,
                type=rel_type,
                source="excel_import",
                confidence=Confidence.DIRECT,
                derived_from=derived_from,
                last_updated=now,
            )
        )

    def _rows(self, ws: Worksheet, columns: list[str]):
        """Yield (row_number, {column: trimmed_value}); skips blank-name rows."""
        rows = ws.iter_rows(values_only=True)
        header = [str(v).strip() if v is not None else "" for v in next(rows)]
        index = {name: i for i, name in enumerate(header)}
        for row_number, raw in enumerate(rows, start=2):
            def get(col: str) -> str:
                i = index[col]
                value = raw[i] if i < len(raw) else None
                return str(value).strip() if value is not None else ""

            if not get(columns[0]):  # blank name -> skip
                continue
            yield row_number, {col: get(col) for col in columns}

    # -- per-sheet parsing ------------------------------------------------
    def _parse_code_sheet(
        self, ws, sheet, label, registry, direct, seen, now,
    ) -> None:
        columns = REQUIRED_COLUMNS[sheet]
        name_col = columns[0]
        for row_number, row in self._rows(ws, columns):
            name = row[name_col]
            self._ensure_node(
                registry, name, label, now, row["Description"], row["BusinessMeaning"]
            )
            derived = f"{sheet}!{row_number}"
            for target in _split(row["Reads"]):
                self._ensure_node(registry, target, "Table", now)
                self._add_edge(direct, seen, name, RelationshipType.READS, target, derived, now)
            for target in _split(row["Writes"]):
                self._ensure_node(registry, target, "Table", now)
                self._add_edge(direct, seen, name, RelationshipType.WRITES, target, derived, now)
            for target in _split(row["Calls"]):
                self._ensure_node(registry, target, "FunctionModule", now)
                self._add_edge(direct, seen, name, RelationshipType.CALLS, target, derived, now)

    def _parse_badis(self, ws, registry, direct, seen, now) -> None:
        for row_number, row in self._rows(ws, REQUIRED_COLUMNS["BAdIs"]):
            name = row["BAdI"]
            self._ensure_node(
                registry, name, "BAdI", now, row["Description"], row["BusinessMeaning"]
            )
            implementer = row["ImplementedBy"]
            if implementer:
                self._ensure_node(registry, implementer, "Program", now)
                # Direction: Program IMPLEMENTS BAdI.
                self._add_edge(
                    direct, seen, implementer, RelationshipType.IMPLEMENTS, name,
                    f"BAdIs!{row_number}", now,
                )

    def _parse_jobs(self, ws, registry, direct, seen, now) -> None:
        for row_number, row in self._rows(ws, REQUIRED_COLUMNS["Jobs"]):
            name = row["Job"]
            self._ensure_node(
                registry, name, "Job", now, row["Description"], row["BusinessMeaning"]
            )
            for target in _split(row["Executes"]):
                self._ensure_node(registry, target, "Program", now)
                self._add_edge(
                    direct, seen, name, RelationshipType.EXECUTES, target,
                    f"Jobs!{row_number}", now,
                )

    def _parse_transports(self, ws, registry, direct, seen, now) -> None:
        for row_number, row in self._rows(ws, REQUIRED_COLUMNS["Transports"]):
            name = row["Transport"]
            # Transports have no BusinessMeaning column.
            self._ensure_node(registry, name, "Transport", now, row["Description"])
            for target in _split(row["Changed"]):
                # Per docs/excel-schema.md, Changed targets pre-exist elsewhere; if
                # one is genuinely new we fall back to Table (its type is unknowable).
                if target not in registry:
                    self._ensure_node(registry, target, "Table", now)
                self._add_edge(
                    direct, seen, name, RelationshipType.CHANGED, target,
                    f"Transports!{row_number}", now,
                )
